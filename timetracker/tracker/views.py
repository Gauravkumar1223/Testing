from django.shortcuts import render, redirect
from django.http import JsonResponse , HttpResponse
import json
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import TimeEntry, CustomUser
from django.urls import reverse
from datetime import datetime,time, timedelta
import calendar
from django.db.models import F
from django.db import transaction
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
import openpyxl
import platform
import io
import zipfile
from PIL import Image as PILImage
# from openpyxl import Workbook




User = get_user_model()


@login_required
def superuser_dashboard(request):
    users = CustomUser.objects.all()
    return render(request, 'superuser_dashboard.html', {'users': users})

def post_login_redirect(request):
    if request.user.is_superuser:
        return redirect('superuser_dashboard')
    else:
        return redirect('home')

def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return post_login_redirect(request)
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'login.html')


def user_logout(request):
    logout(request)
    return redirect('login')

def forgot_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        new_password = request.POST.get('new_password')
        
        try:
            user = User.objects.get(username=username)
            user.set_password(new_password)
            user.save()
            messages.success(request, 'Password updated successfully. You can now log in.')
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
            return render(request, 'forgot_password.html')

    return render(request, 'forgot_password.html')


# Helper function to get the dates of a specific week
def get_week_dates(year, week):
    first_day = datetime.strptime(f'{year}-W{week}-1', "%Y-W%W-%w").date()
    return [first_day + timedelta(days=i) for i in range(7)]

# Helper function to get the weeks for a specific month
def get_month_weeks(year, month):

    first_day = datetime(year, month, 1)
    
    # Get the last day of the month
    if month == 12:
        # For December, next month is January of the next year
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        # For other months, just get the last day of the next month
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    
    start_weekday = first_day.weekday()  # Monday is 0, Sunday is 6
    days_in_month = (last_day - first_day).days + 1

    # Create a list to store weeks
    weeks = []
    current_week = [None] * start_weekday  
    for day in range(1, days_in_month + 1):
        current_week.append(datetime(year, month, day).date())

        if len(current_week) == 7:
            weeks.append(current_week)
            current_week = []

    if current_week:
        weeks.append(current_week + [None] * (7 - len(current_week)))  # Pad with None

    return weeks

# Display a dynamic month
@login_required
def home(request):
    # Use URL parameters if available, otherwise default to today
    year = int(request.GET.get('year', datetime.today().year))
    month = int(request.GET.get('month', datetime.today().month))

    # Handle wrap-around logic for previous and next months
    prev_month = month - 1 if month > 1 else 12
    prev_month_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_month_year = year if month < 12 else year + 1
    month_weeks = get_month_weeks(year, month)

    # Fetch time entries for the current user and month
    user_entries = TimeEntry.objects.filter(user=request.user, date__year=year, date__month=month)
    time_entries = {entry.date: entry for entry in user_entries}

    month_name = calendar.month_name[month]

    context = {
        'month_weeks': month_weeks,
        'time_entries': time_entries,
        'year': year,
        'month': month,
        'prev_month': prev_month,
        'prev_month_year': prev_month_year,
        'next_month': next_month,
        'next_month_year': next_month_year,
        'month_name': month_name,
    }

    return render(request, 'home.html', context)


@require_POST
@login_required
def auto_fill(request):
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))

    _, last_day = calendar.monthrange(year, month)

    for day in range(1, last_day + 1):
        date_obj = datetime(year, month, day).date()
        if date_obj.weekday() < 5: 
            TimeEntry.objects.get_or_create(
                user=request.user,
                date=date_obj,
                defaults={'hours': 8, 'day_type': 'work'}
            )

    return redirect(f"{reverse('home')}?year={year}&month={month}")


# AJAX view to update time entry
def ajax_update_entry(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        date = request.POST.get('date')
        hours = request.POST.get('hours')
        description = request.POST.get('description')
        day_type = request.POST.get('day_type')

        # Ensure all necessary data is provided
        if not all([date, hours, day_type]):
            return JsonResponse({'status': 'error', 'message': 'Missing required fields'})

        try:
            # Parse the date and convert it to a datetime object
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid date format'})

        try:
            with transaction.atomic():
                entry, created = TimeEntry.objects.select_for_update().get_or_create(
                    user=request.user,
                    date=date_obj,
                    defaults={'hours': hours, 'description': description, 'day_type': day_type}
                )

                previous_day_type = entry.day_type if not created else None

                # Handle vacation days accounting
                if created:
                    if day_type == 'vacation':
                        if request.user.vacation_days_old > 0:
                            request.user.vacation_days_old = F('vacation_days_old') - 1
                            request.user.save()
                        elif request.user.vacation_days > 0:
                            request.user.vacation_days = F('vacation_days') - 1
                            request.user.save()
                        else:
                            return JsonResponse({'status': 'error', 'message': 'No vacation days left'})

                else:
                    if previous_day_type != day_type:
                        # Vacation day was removed
                        if previous_day_type == 'vacation' and day_type != 'vacation':
                            if request.user.vacation_days_old > 0 :
                                request.user.vacation_days_old = F('vacation_days_old') + 1
                            elif request.user.vacation_days < 20 :
                                request.user.vacation_days = F('vacation_days') + 1
                            else:
                                return  JsonResponse({'status': 'error', 'message': 'You can"t revoke old vacation days, talk to the system admin!'})
                            request.user.save()

                        # Vacation day was added
                        elif previous_day_type != 'vacation' and day_type == 'vacation':
                            if request.user.vacation_days_old > 0:
                                request.user.vacation_days_old = F('vacation_days_old') - 1
                            elif request.user.vacation_days > 0:
                                request.user.vacation_days = F('vacation_days') - 1
                            else:
                                return JsonResponse({'status': 'error', 'message': 'No vacation days left'})
                            request.user.save()

                entry.hours = hours
                entry.description = description
                entry.day_type = day_type
                entry.save()

            return JsonResponse({'status': 'success'})

        except Exception as e:
            print(f"Error saving time entry: {str(e)}")
            return JsonResponse({'status': 'error', 'message': 'An error occurred while saving the entry'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


@csrf_exempt
def save_user_data(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        user_id = data.get('id')

        if user_id:
            user = User.objects.get(id=user_id)
        else:
            user = User()

        user.first_name = data['first_name']
        user.last_name = data['last_name']
        user.username = data['username']
        user.email = data['email']
        user.position = data['position']
        user.team_lead = data['team_lead']
        user.project_manager = data['project_manager']
        user.project = data['project']
        user.vacation_days = data['vacation_days']
        user.save()
        return JsonResponse({'status': 'success'})

    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def superuser_dashboard(request):
    users = CustomUser.objects.all()
    year = int(request.GET.get('year', datetime.today().year))
    month = int(request.GET.get('month', datetime.today().month))

    prev_month = month - 1 if month > 1 else 12
    prev_month_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_month_year = year if month < 12 else year + 1
    month_name = calendar.month_name[month]

    context = {
        'users': users,
        'year': year,
        'month': month,
        'month_name': month_name,
        'prev_month': prev_month,
        'prev_month_year': prev_month_year,
        'next_month': next_month,
        'next_month_year': next_month_year,
    }

    return render(request, 'superuser_dashboard.html', context)

@login_required
def export_user_excel(request):
    user_id = request.GET.get("user_id")
    year = int(request.GET.get("year"))
    month = int(request.GET.get("month"))

    user = CustomUser.objects.get(id=user_id)
    month_name_str = calendar.month_name[month]
    days_in_month = calendar.monthrange(year, month)[1]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

    # --- Header Section ---
    ws.merge_cells('A1:B1')
    ws.merge_cells('C1:D1')
    ws['A1'] = 'Timesheet -'
    ws['C1'] = f"{month_name_str} {year}"
    for cell in ['A1', 'C1']:
        ws[cell].font = header_font
        ws[cell].alignment = center_alignment
        ws[cell].fill = blue_fill

    ws['A2'] = 'Resource Name:'
    ws['B2'] = f"{user.first_name} {user.last_name}"
    ws['C2'] = 'Position:'
    ws['D2'] = user.position

    ws['A3'] = 'Team Lead:'
    ws['B3'] = user.team_lead
    ws['C3'] = 'Client:'
    ws['D3'] = user.project

    ws['A4'] = 'Project Manager:'
    ws['B4'] = user.project_manager
    ws['C4'] = 'Project:'
    ws['D4'] = user.project

    for cell in ['A2', 'A3', 'A4', 'C2', 'C3', 'C4']:
        ws[cell].font = bold_font

    column_widths = {'A': 28, 'B': 12, 'C': 12, 'D': 18, 'E': 18, 'F': 12}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # --- Table Header for Daily Log ---
    start_row = 6
    headers = ['Date', 'Start Time', 'End Time', 'Billable Hrs', 'Non-billable Hrs', 'Total Hrs']
    for i, header in enumerate(headers):
        col = chr(65 + i)
        ws[f"{col}{start_row}"] = header
        ws[f"{col}{start_row}"].font = header_font
        ws[f"{col}{start_row}"].alignment = center_alignment
        ws[f"{col}{start_row}"].fill = blue_fill

    # --- Daily Log ---
    current_row = start_row + 1
    for day in range(1, days_in_month + 1):
        date_obj = datetime(year, month, day)
        day_str = date_obj.strftime("%A, %B %#d, %Y") if platform.system() == "Windows" else date_obj.strftime("%A, %B %-d, %Y")

        entries = TimeEntry.objects.filter(user=user, date=date_obj.date())
        total_hours = sum([entry.hours for entry in entries])

        is_zero_day = total_hours <= 0
        start_time = time(0, 0)
        end_time = time(0, 0)
        billable_time = time(0, 0)
        non_billable_time = time(0, 0)
        total_time_display = "-"

        if not is_zero_day:
            start_time = time(8, 0)
            total_minutes = int(total_hours * 60)
            end_time = (datetime.combine(datetime.today(), start_time) + timedelta(minutes=total_minutes + 60)).time() 
            billable_time = (datetime.combine(datetime.today(), time(0, 0)) + timedelta(minutes=total_minutes + 60)).time()
            non_billable_time = time(1, 0)
            total_time_display = float(total_hours)

        ws[f"A{current_row}"] = day_str
        ws[f"B{current_row}"] = start_time
        ws[f"C{current_row}"] = end_time
        ws[f"D{current_row}"] = billable_time
        ws[f"E{current_row}"] = non_billable_time
        ws[f"F{current_row}"] = total_time_display if not is_zero_day else None

        # Apply formatting
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f"{col}{current_row}"]
            cell.alignment = center_alignment
            if col in ['B', 'C', 'D', 'E'] and not is_zero_day:
                cell.number_format = 'h:mm'
            elif col == 'F' and not is_zero_day:
                cell.number_format = '0.00'
            if is_zero_day:
                cell.fill = grey_fill

        current_row += 1

    # --- Total Hrs Row ---
    total_row = current_row
    ws[f"E{total_row}"] = "Total Hrs:"
    ws[f"E{total_row}"].font = Font(bold=True)
    ws[f"E{total_row}"].alignment = center_alignment

    ws[f"F{total_row}"] = f"=SUM(F{start_row + 1}:F{current_row - 1})"
    ws[f"F{total_row}"].font = Font(bold=True)
    ws[f"F{total_row}"].alignment = center_alignment
    ws[f"F{total_row}"].number_format = '0.00'

    # --- Signature ---
    signature_row = total_row + 2
    ws[f"A{signature_row}"] = 'Resource Signature:'
    ws[f"A{signature_row}"].font = bold_font

    if user.signature:
        try:
            image_stream = io.BytesIO(user.signature)
            pil_img = PILImage.open(image_stream)

            temp_img_stream = io.BytesIO()
            pil_img.save(temp_img_stream, format='PNG')
            temp_img_stream.seek(0)

            img = ExcelImage(temp_img_stream)
            img.width = 100
            img.height = 30
            img.anchor = f"B{signature_row}"
            ws.add_image(img)
        except Exception as e:
            print(f"Error inserting signature for user {user.username}: {e}")
            ws[f"B{signature_row}"] = ""

    # Save Excel file to response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Timesheet_RA_{user.first_name}_{user.last_name}_{month_name_str}_{year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_all_users_excel(request):
    year = int(request.GET.get("year"))
    month = int(request.GET.get("month"))
    month_name_str = calendar.month_name[month]
    days_in_month = calendar.monthrange(year, month)[1]

    users = CustomUser.objects.all()
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for user in users:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Timesheet"

            # Styles
            bold_font = Font(bold=True)
            header_font = Font(bold=True, color="FFFFFF", size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

            # Header
            ws.merge_cells('A1:B1')
            ws.merge_cells('C1:D1')
            ws['A1'] = 'Timesheet -'
            ws['C1'] = f"{month_name_str} {year}"
            for cell in ['A1', 'C1']:
                ws[cell].font = header_font
                ws[cell].alignment = center_alignment
                ws[cell].fill = blue_fill

            ws['A2'] = 'Resource Name:'
            ws['B2'] = f"{user.first_name} {user.last_name}"
            ws['C2'] = 'Position:'
            ws['D2'] = user.position

            ws['A3'] = 'Team Lead:'
            ws['B3'] = user.team_lead
            ws['C3'] = 'Client:'
            ws['D3'] = user.project

            ws['A4'] = 'Project Manager:'
            ws['B4'] = user.project_manager
            ws['C4'] = 'Project:'
            ws['D4'] = user.project

            for cell in ['A2', 'A3', 'A4', 'C2', 'C3', 'C4']:
                ws[cell].font = bold_font

            column_widths = {'A': 28, 'B': 12, 'C': 12, 'D': 18, 'E': 18, 'F': 12}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            start_row = 6
            headers = ['Date', 'Start Time', 'End Time', 'Billable Hrs', 'Non-billable Hrs', 'Total Hrs']
            for i, header in enumerate(headers):
                col = chr(65 + i)
                ws[f"{col}{start_row}"] = header
                ws[f"{col}{start_row}"].font = header_font
                ws[f"{col}{start_row}"].alignment = center_alignment
                ws[f"{col}{start_row}"].fill = blue_fill

            # Daily Log
            current_row = start_row + 1
            for day in range(1, days_in_month + 1):
                date_obj = datetime(year, month, day)
                day_str = date_obj.strftime("%A, %B %#d, %Y") if platform.system() == "Windows" else date_obj.strftime("%A, %B %-d, %Y")

                entries = TimeEntry.objects.filter(user=user, date=date_obj.date())
                total_hours = sum(entry.hours for entry in entries)

                is_zero_day = total_hours <= 0
                start_time = time(0, 0)
                end_time = time(0, 0)
                billable_time = time(0, 0)
                non_billable_time = time(0, 0)
                total_time_display = "-"

                if not is_zero_day:
                    start_time = time(8, 0)
                    total_minutes = int(total_hours * 60)
                    end_time = (datetime.combine(datetime.today(), start_time) + timedelta(minutes=total_minutes + 60)).time() 
                    billable_time = (datetime.combine(datetime.today(), time(0, 0)) + timedelta(minutes=total_minutes + 60)).time()
                    non_billable_time = time(1, 0)
                    total_time_display = float(total_hours)

                ws[f"A{current_row}"] = day_str
                ws[f"B{current_row}"] = start_time
                ws[f"C{current_row}"] = end_time
                ws[f"D{current_row}"] = billable_time
                ws[f"E{current_row}"] = non_billable_time
                ws[f"F{current_row}"] = total_time_display if not is_zero_day else None

                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    cell = ws[f"{col}{current_row}"]
                    cell.alignment = center_alignment
                    if col in ['B', 'C', 'D', 'E'] and not is_zero_day:
                        cell.number_format = 'h:mm'
                    elif col == 'F' and not is_zero_day:
                        cell.number_format = '0.00'
                    if is_zero_day:
                        cell.fill = grey_fill

                current_row += 1

            # Total Hours
            total_row = current_row
            ws[f"E{total_row}"] = "Total Hrs:"
            ws[f"E{total_row}"].font = Font(bold=True)
            ws[f"E{total_row}"].alignment = center_alignment
            ws[f"F{total_row}"] = f"=SUM(F{start_row + 1}:F{current_row - 1})"
            ws[f"F{total_row}"].font = Font(bold=True)
            ws[f"F{total_row}"].alignment = center_alignment
            ws[f"F{total_row}"].number_format = '0.00'

            signature_row = total_row + 2

            ws[f"A{signature_row}"] = 'Resource Signature:'
            ws[f"A{signature_row}"].font = bold_font

            if user.signature:
                try:
                    image_stream = io.BytesIO(user.signature)
                    pil_img = PILImage.open(image_stream)

                    temp_img_stream = io.BytesIO()
                    pil_img.save(temp_img_stream, format='PNG')
                    temp_img_stream.seek(0)

                    img = ExcelImage(temp_img_stream)
                    img.width = 100
                    img.height = 30
                    img.anchor = f"B{signature_row}"
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error inserting signature for user {user.username}: {e}")
                    ws[f"B{signature_row}"] = ""

            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)
            filename = f"Timesheet_RA_{user.first_name}_{user.last_name}_{month_name_str}_{year}.xlsx"
            zip_file.writestr(filename, file_stream.read())

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="Timesheets_RA_{month_name_str}_{year}.zip"'
    return response




